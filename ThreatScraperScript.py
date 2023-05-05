import requests
import pandas as pd
import datetime
import schedule
import time
from openpyxl import load_workbook

# API key from VirusTotal
api_key = 'Enter your VirusTotal API here'

# SHA-256 hash of the file you want to check
sha256_hash = 'Enter your SHA1 / SHA256 / MD5 here'

# URL endpoint for VirusTotal API
url = f'https://www.virustotal.com/vtapi/v2/file/report?apikey={api_key}&resource={sha256_hash}'

# Load existing workbook. Replace the line below with your current Excel file that you want to dump the results to.
filename = r'C:\VirusTotal\Ancienttotem.xlsx'
wb = load_workbook(filename)

def get_report():
    # Send API request
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Convert the response to a pandas DataFrame
        df = pd.json_normalize(response.json())

        # Select the worksheet where you want to add the new report
        ws = wb.active

        # Find the next empty row
        row = ws.max_row + 1

        # Convert the DataFrame to a list of lists
        rows = df.values.tolist()

        # Write the rows to the worksheet
        for row_data in rows:
            ws.append(row_data)

        # Save the workbook to the Excel file
        wb.save(filename)

        print(f'Report saved to C:\VirusTotal\Ancienttotem.xlsx')
    else:
        # Print error message
        print(f'Error: {response.status_code} - {response.reason}')

# Schedule the job to run every two hours every day
schedule.every().day.at('02:26').do(get_report)
schedule.every().day.at('04:26').do(get_report)
schedule.every().day.at('06:26').do(get_report)
schedule.every().day.at('08:26').do(get_report)
schedule.every().day.at('10:26').do(get_report)
schedule.every().day.at('12:26').do(get_report)
schedule.every().day.at('14:26').do(get_report)
schedule.every().day.at('16:26').do(get_report)
schedule.every().day.at('18:26').do(get_report)
schedule.every().day.at('20:26').do(get_report)
schedule.every().day.at('22:26').do(get_report)
schedule.every().day.at('00:26').do(get_report)

# Run the scheduler
while True:
    schedule.run_pending()
    time.sleep(1)
