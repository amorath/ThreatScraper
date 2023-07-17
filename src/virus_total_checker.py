import time
from tkinter import ttk
import tkinter as tk
import matplotlib.pyplot as plt
import pandas as pd
import requests
from openpyxl import load_workbook
import os
import hashlib

class VirusTotalChecker:
    def __init__(self, api_key, hash_value, hash_type, filename, console_output, master):
        self.api_key = api_key
        self.hash_value = hash_value
        self.hash_type = hash_type
        self.url = f'https://www.virustotal.com/vtapi/v2/file/report?apikey={api_key}&resource={hash_value}'
        self.filename = filename
        self.console_output = console_output
        self.figure, self.ax = plt.subplots()  # Create figure and axes for the line graph
        self.figure.suptitle("Detection Timeline", fontsize=14)  # Set window title for the line graph
        self.figure2, self.ax2 = plt.subplots()  # Create a separate figure and axes for the pie chart
        self.figure2.suptitle("Detection Results", fontsize=14)  # Set window title for the pie chart
        self.line, = self.ax.plot([], [], label='Malware Detections')  # Create an empty line
        self.line_total, = self.ax.plot([], [], label='AV Utilization')  # Create a second line for total scans
        self.master = master  # Add master to the VirusTotalChecker class
        self.top = None  # Add a member to keep track of the Toplevel window

    def submit_file(self, file_path):
        url = 'https://www.virustotal.com/vtapi/v2/file/scan'

        params = {'apikey': self.api_key}

        files = {'file': (os.path.basename(file_path), open(file_path, 'rb'))}

        response = requests.post(url, files=files, params=params)
        
        if response.status_code == 200:
            self.console_output.write(f'File submitted for analysis: {file_path}\n')
            self.calculate_hashes(file_path)
        else:
            self.console_output.write(f'Error submitting file: {response.status_code} - {response.reason}\n')

    def calculate_hashes(self, file_path):
        BUF_SIZE = 65536  # read file in 64kb chunks

        md5 = hashlib.md5()
        sha1 = hashlib.sha1()
        sha256 = hashlib.sha256()

        with open(file_path, 'rb') as f:
            while True:
                data = f.read(BUF_SIZE)
                if not data:
                    break

                md5.update(data)
                sha1.update(data)
                sha256.update(data)

        self.console_output.write(f'MD5: {md5.hexdigest()}\n')
        self.console_output.write(f'SHA1: {sha1.hexdigest()}\n')
        self.console_output.write(f'SHA256: {sha256.hexdigest()}\n')

    def rescan_hash(self):
        # API endpoint for rescan
        rescan_url = f'https://www.virustotal.com/api/v3/files/{self.hash_value}/analyse'
        headers = {
            "accept": "application/json",
            "x-apikey": self.api_key
        }

        response = requests.post(rescan_url, headers=headers)

        if response.status_code == 200:
            self.console_output.write(f'Rescan started for {self.hash_value}\n')
            time.sleep(10)  # wait for 10 seconds before pulling the report
            self.check_virustotal()  # pull the report after rescan
        else:
            self.console_output.write(f'Error starting rescan: {response.status_code} - {response.reason}\n')     

    def process_virustotal_results(self, json_response):
        """
        Process the VirusTotal API JSON response to separate the results by Anti Virus
        programs that detected malware and those that didn't.
        """
        scans = json_response.get('scans', {})
        positive_results = {av: details for av, details in scans.items() if details.get('detected', False)}
        negative_results = {av: details for av, details in scans.items() if not details.get('detected', False)}

        # Convert the results into pandas DataFrame
        self.positive_df = pd.DataFrame.from_dict(positive_results, orient='index')
        self.negative_df = pd.DataFrame.from_dict(negative_results, orient='index')

    def show_results(self):
        """
        Show the results in a separate tkinter window with two Treeviews.
        """
        # Create a new top level window
        if self.top is None:  # If the Toplevel window doesn't exist yet, create one
            self.top = tk.Toplevel(self.master)
            self.top.iconbitmap('ThreatScraper.ico')
            self.top.title("Scan Results")
            self.top.geometry("")  # Set the initial size of the window (optional)
            self.top.minsize(200, 200)  # Set the minimum size of the window
            self.top.maxsize(2000, 2000)  # Set the maximum size of the window
        else:  # If it does exist, clear it
            for widget in self.top.winfo_children():
                widget.destroy()

        # Create a PanedWindow to hold the treeviews
        paned = tk.PanedWindow(self.top, orient='horizontal')
        paned.pack(fill='both', expand=True)

        # Create two treeviews
        positive_tree = ttk.Treeview(paned)
        negative_tree = ttk.Treeview(paned)

        # Add columns to the treeviews
        positive_tree['columns'] = list(self.positive_df.columns)
        negative_tree['columns'] = list(self.negative_df.columns)

        # Sort DataFrame by index
        self.positive_df.sort_index(inplace=True)
        self.negative_df.sort_index(inplace=True)

        # Add data to the treeviews
        for index, row in self.positive_df.iterrows():
            positive_tree.insert('', 'end', text=index, values=list(row))

        for index, row in self.negative_df.iterrows():
            negative_tree.insert('', 'end', text=index, values=list(row))

        # Add the treeviews to the PanedWindow
        paned.add(positive_tree)
        paned.add(negative_tree)

    def check_virustotal(self):
        # Send API request
        response = requests.get(self.url)

        # Check if the request was successful
        if response.status_code == 200:

            # Call the new method to process results
            self.process_virustotal_results(response.json())

            # Call the new method to show results
            self.show_results()

            # Convert the response to a pandas DataFrame
            df = pd.json_normalize(response.json())

            wb = load_workbook(self.filename)
            ws = wb.active

            # Find the next empty row
            ws.max_row + 1

            # Get the existing column headers with whitespaces
            existing_headers = [str(header.value).strip() if isinstance(header.value, str) else str(header.value) for header in ws[1]]

            # Update the column headers without whitespaces
            headers = [header.strip() for header in existing_headers]

            # Convert the DataFrame to a list of lists
            rows = df.values.tolist()

            # Write the rows to the worksheet
            for row_data in rows:
                ws.append(row_data)

            # Update the column headers if they have changed
            if existing_headers != headers:
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)

            # Save the workbook to the Excel file
            wb.save(self.filename)

            # Print success message
            self.console_output.write(f'Report saved to {self.filename}\n')
        else:
            # Print error message
            self.console_output.write(f'Error: {response.status_code} - {response.reason}\n')

        # Update graph
        try:
            df = pd.read_excel(self.filename)

            # Strip the leading and trailing whitespaces from column headers
            df.columns = [str(column).strip() for column in df.columns]

            # Specify the correct column name that contains the numeric values
            target_column = 'positives'  # Replace with the correct column name

            # Ensure that the target column exists and is numeric
            if target_column in df.columns and pd.api.types.is_numeric_dtype(df[target_column]):
                self.load_data(self.filename)  # Load data from file
                self.update_graph()  # Update the line graph
                self.update_pie_chart()  # Update the pie chart
            else:
                self.console_output.write(
                    f'Error: column "{target_column}" is not present or '
                    f'not numeric in {self.filename}\n'
                )

        except Exception as e:
            self.console_output.write(f'Error updating graph: {str(e)}\n')

    def load_data(self, filename):
        try:
            df = pd.read_excel(filename)
            self.ydata = df['positives'].tolist()
            self.ydata_total = df['total'].tolist()  # Get data for 'total' column
            self.xdata = list(range(1, len(self.ydata) + 1))  # Update x-axis values from 1 to the length of ydata
            self.line.set_data([], [])  # Clear previous data
            self.line_total.set_data([], [])  # Clear previous data for 'total' line
        except Exception as e:
            self.console_output.write(f'Error loading data: {str(e)}\n')

        # New: Load data for pie chart
        self.positives = df['positives'].tolist()[-1]
        self.negatives = df['total'].tolist()[-1] - self.positives

    def update_graph(self):
        self.line.set_data(self.xdata, self.ydata)  # Update the line data
        self.line_total.set_data(self.xdata, self.ydata_total)  # Update the 'total' line data
        self.ax.relim()  # Recalculate limits
        self.ax.autoscale_view(True, True, True)  # Rescale the view
        plt.figure(self.figure.number)  # Select the figure for the line graph
        plt.xlabel('Scan Number')
        plt.ylabel('Malware Detections / AV Utilization')
        plt.legend()
        plt.draw()
        plt.pause(0.001)
        plt.show(block=False)  # Display the figure without blocking

    def start(self):
        self.load_data(self.filename)
        self.update_graph()
        self.update_pie_chart()  # Update the pie chart
        plt.show(block=False)  # Display both figures without blocking

    def update_pie_chart(self):
        labels = ['Positive Findings', 'Negative Findings']
        sizes = [self.positives, self.negatives]
        self.ax2.clear()  # Clear previous data
        self.ax2.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
        plt.figure(self.figure2.number)  # Select the figure for the pie chart
        plt.axis('equal')  # Equal aspect ratio ensures the pie chart is circular
        plt.show(block=False)  # Display the figure without blocking
