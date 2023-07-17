import sys
import threading
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import requests
import pandas as pd
import datetime
import schedule
import time
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from multiprocessing import Process
import json
import os

class ConsoleOutput:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

    def set_text_widget_state(self, state):
        self.text_widget.configure(state=state)


class VirusTotalChecker:
    def __init__(self, api_key, hash_value, hash_type, filename, console_output, master):
        self.api_key = api_key
        self.hash_value = hash_value
        self.hash_type = hash_type
        self.url = f'https://www.virustotal.com/vtapi/v2/file/report?apikey={api_key}&resource={hash_value}'
        self.filename = filename
        self.console_output = console_output
        self.figure, self.ax = plt.subplots()  # Create figure and axes
        self.line, = self.ax.plot([], [], label='Malware Detections')  # Create an empty line
        self.line_total, = self.ax.plot([], [], label='AV Utilization')  # Create a second line for total scans
        self.master = master  # Add master to the VirusTotalChecker class
        self.top = None  # Add a member to keep track of the Toplevel window

    def rescan_hash(self):
        # API endpoint for rescan
        rescan_url = f'https://www.virustotal.com/api/v3/files/{self.sha256_hash}/analyse'
        headers = {
            "x-apikey": self.api_key
        }

        response = requests.post(rescan_url, headers=headers)

        if response.status_code == 200:
            self.console_output.write(f'Rescan started for {self.sha256_hash}\n')
            time.sleep(10)  # wait for 10 seconds before pulling the report
            self.check_virustotal()  # pull the report after rescan
        else:
            self.console_output.write(f'Error starting rescan: {response.status_code} - {response.reason}\n')    

    def process_virustotal_results(self, json_response):
        """
        Process the VirusTotal API JSON response to separate the results by Anti Virus
        programs that detected malware and those that didn't.
        """
        scans = json_response.get('scans', {})
        positive_results = {av: details for av, details in scans.items() if details.get('detected', False)}
        negative_results = {av: details for av, details in scans.items() if not details.get('detected', False)}

        # Convert the results into pandas DataFrame
        self.positive_df = pd.DataFrame.from_dict(positive_results, orient='index')
        self.negative_df = pd.DataFrame.from_dict(negative_results, orient='index')

    def show_results(self):
        """
        Show the results in a separate tkinter window with two Treeviews.
        """
        # Create a new top level window
        if self.top is None:  # If the Toplevel window doesn't exist yet, create one
            self.top = tk.Toplevel(self.master)
            self.top.iconbitmap('ThreatScraper.ico')
            self.top.title("Scan Results")
            self.top.geometry("")  # Set the initial size of the window (optional)
            self.top.minsize(200, 200)  # Set the minimum size of the window
            self.top.maxsize(2000, 2000)  # Set the maximum size of the window
        else:  # If it does exist, clear it
            for widget in self.top.winfo_children():
                widget.destroy()

        # Create a PanedWindow to hold the treeviews
        paned = tk.PanedWindow(self.top, orient='horizontal')
        paned.pack(fill='both', expand=True)

        # Create two treeviews
        positive_tree = ttk.Treeview(paned)
        negative_tree = ttk.Treeview(paned)

        # Add columns to the treeviews
        positive_tree['columns'] = list(self.positive_df.columns)
        negative_tree['columns'] = list(self.negative_df.columns)

        # Add data to the treeviews
        for index, row in self.positive_df.iterrows():
            positive_tree.insert('', 'end', text=index, values=list(row))

        for index, row in self.negative_df.iterrows():
            negative_tree.insert('', 'end', text=index, values=list(row))

        # Add the treeviews to the PanedWindow
        paned.add(positive_tree)
        paned.add(negative_tree)


    def check_virustotal(self):
        # Send API request
        response = requests.get(self.url)

        # Check if the request was successful
        if response.status_code == 200:

            # Call the new method to process results
            self.process_virustotal_results(response.json())

            # Call the new method to show results
            self.show_results()

            # Convert the response to a pandas DataFrame
            df = pd.json_normalize(response.json())

            wb = load_workbook(self.filename)
            ws = wb.active

            # Find the next empty row
            ws.max_row + 1

            # Get the existing column headers with whitespaces
            existing_headers = [str(header.value).strip() if isinstance(header.value, str) else str(header.value) for header in ws[1]]

            # Update the column headers without whitespaces
            headers = [header.strip() for header in existing_headers]

            # Convert the DataFrame to a list of lists
            rows = df.values.tolist()

            # Write the rows to the worksheet
            for row_data in rows:
                ws.append(row_data)

            # Update the column headers if they have changed
            if existing_headers != headers:
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)

            # Save the workbook to the Excel file
            wb.save(self.filename)

            # Print success message
            self.console_output.write(f'Report saved to {self.filename}\n')
        else:
            # Print error message
            self.console_output.write(f'Error: {response.status_code} - {response.reason}\n')

        # Update graph
        try:
            df = pd.read_excel(self.filename)

            # Strip the leading and trailing whitespaces from column headers
            df.columns = [str(column).strip() for column in df.columns]

            # Specify the correct column name that contains the numeric values
            target_column = 'positives'  # Replace with the correct column name

            # Ensure that the target column exists and is numeric
            if target_column in df.columns and pd.api.types.is_numeric_dtype(
                df[target_column]
            ):
                df[target_column].iloc[-1]
                self.load_data(self.filename)  # Load data from file
                self.update_graph()  # Update the graph
            else:
                self.console_output.write(
                    f'Error: column "{target_column}" is not present or '
                    f'not numeric in {self.filename}\n'
                )

        except Exception as e:
            self.console_output.write(f'Error updating graph: {str(e)}\n')

    def load_data(self, filename):
        try:
            df = pd.read_excel(filename)
            self.ydata = df['positives'].tolist()
            self.ydata_total = df['total'].tolist()  # Get data for 'total' column
            self.xdata = list(range(1, len(self.ydata) + 1))  # Update x-axis values from 1 to the length of ydata
            self.line.set_data([], [])  # Clear previous data
            self.line_total.set_data([], [])  # Clear previous data for 'total' line
        except Exception as e:
            self.console_output.write(f'Error loading data: {str(e)}\n')


    def update_graph(self):
        self.line.set_data(self.xdata, self.ydata)  # Update the line data
        self.line_total.set_data(self.xdata, self.ydata_total)  # Update the 'total' line data
        self.ax.relim()  # Recalculate limits
        self.ax.autoscale_view(True, True, True)  # Rescale the view
        plt.xlabel('Scan Number')
        plt.ylabel('Malware Detections / AV Utilization')
        plt.legend()
        plt.draw()
        plt.pause(0.001)

    def start(self):
        self.load_data(self.filename)
        self.update_graph()

class App:
    def __init__(self, master):
        self.master = master
        master.title('ThreatScraper')
        self.threat_checker = None  # Initialize the VirusTotalChecker object as None

        # Create API key entry
        self.api_key_label = tk.Label(master, text='API Key:')
        self.api_key_label.grid(row=0, column=0)
        self.api_key_entry = tk.Entry(master, width=50)
        self.api_key_entry.grid(row=0, column=1)

        # Create hash value entry
        self.hash_value_label = tk.Label(master, text='Hash Value:')
        self.hash_value_label.grid(row=1, column=0)
        self.hash_value_entry = tk.Entry(master, width=50)
        self.hash_value_entry.grid(row=1, column=1)
        
        # Create hash type selection
        self.hash_type = tk.StringVar(master)  # create a tkinter string variable
        self.hash_type.set("SHA-256")  # set default value
        self.hash_types = ["MD5", "SHA-1", "SHA-256"]  # list of hash types
        self.hash_type_option = tk.OptionMenu(master, self.hash_type, *self.hash_types)
        self.hash_type_option.grid(row=1, column=2)

        # Create filename entry
        self.filename_label = tk.Label(master, text='Filename:')
        self.filename_label.grid(row=2, column=0)
        self.filename_entry = tk.Entry(master, width=50)
        self.filename_entry.grid(row=2, column=1)

        # Create time entry
        self.time_label = tk.Label(master, text='Schedule Times (HH:MM) Separated by comma:')
        self.time_label.grid(row=3, column=0)
        self.time_entry = tk.Entry(master, width=50)
        self.time_entry.grid(row=3, column=1)

        # Load saved configurations
        self.config_filename = 'config.json'
        self.load_config()

        # Checkbox for rescanning
        self.rescan_var = tk.BooleanVar()
        self.rescan_checkbox = tk.Checkbutton(master, text="Rescan hash", variable=self.rescan_var)
        self.rescan_checkbox.grid(row=2, column=2, columnspan=2)

        # Create check button
        self.check_button = tk.Button(master, text='Check VirusTotal', command=self.check_virustotal)
        self.check_button.grid(row=5, column=1)

        # Create start button
        self.start_button = tk.Button(master, text='Start Schedule', command=self.start_schedule)
        self.start_button.grid(row=5, column=2)

        # Create stop button
        self.stop_button = tk.Button(master, text='Stop Schedule', command=self.stop_schedule, state='disabled')
        self.stop_button.grid(row=6, column=2)

        # Create console output in main window
        self.console_text = tk.Text(master, state='disabled', height=5)
        self.console_text.grid(row=7, column=0, columnspan=2, sticky='nsew')
        self.console_output = ConsoleOutput(self.console_text)
        self.console_output.set_text_widget_state('normal')
        sys.stdout = self.console_output

        # Configure the grid to make the console text box resize with the window
        master.grid_rowconfigure(7, weight=1)
        master.grid_columnconfigure(0, weight=1)
        master.grid_columnconfigure(1, weight=1)

        # Initialize schedule variables
        self.schedule_thread = None
        self.schedule_running = False

        # Start the graph
        self.threat_graph = None

    def load_config(self):
        # Check if config file exists
        if os.path.exists(self.config_filename):
            with open(self.config_filename, 'r') as f:
                config = json.load(f)
                
                # Load saved API key
                if 'api_key' in config:
                    self.api_key_entry.insert(0, config['api_key'])
                
                # Load saved hash value
                if 'hash_value' in config:
                    self.hash_value_entry.insert(0, config['hash_value'])

                # Load saved hash type
                if 'hash_type' in config:
                    self.hash_type.set(config['hash_type'])
                
                # Load saved filename
                if 'filename' in config:
                    self.filename_entry.insert(0, config['filename'])
                
                # Load saved schedule times
                if 'schedule_times' in config:
                    self.time_entry.insert(0, config['schedule_times'])

    def save_config(self):
        config = {
            'api_key': self.api_key_entry.get(),
            'hash_value': self.hash_value_entry.get(),
            'hash_type': self.hash_type.get(),
            'filename': self.filename_entry.get(),
            'schedule_times': self.time_entry.get(),
        }

        with open(self.config_filename, 'w') as f:
            json.dump(config, f, indent=4)

    def check_virustotal(self):
        api_key = self.api_key_entry.get()
        hash_value = self.hash_value_entry.get()
        hash_type = self.hash_type.get()
        filename = self.filename_entry.get()

        if self.threat_checker is None:
            self.threat_checker = VirusTotalChecker(api_key, hash_value, hash_type, filename, self.console_output, self.master)
        
        if self.rescan_var.get():  # if checkbox is checked
            self.master.after(0, self.threat_checker.rescan_hash)
        else:
            self.master.after(0, self.threat_checker.check_virustotal)

        # Save configurations before running the check
        self.save_config()


    def start_schedule(self):
        # Get schedule times from entry
        schedule_times = self.time_entry.get().split(',')

        # Disable start button and enable stop button
        self.start_button.configure(state='disabled')
        self.stop_button.configure(state='normal')

        # Loop through schedule times and schedule checks
        for schedule_time in schedule_times:
            # Check if time is valid
            try:
                datetime.datetime.strptime(schedule_time.strip(), '%H:%M')
            except ValueError:
                messagebox.showerror(
                    "Invalid Time",
                    f"{schedule_time} is not a valid time. Please enter a "
                    f"valid time in HH:MM format",
                )
                return
            
        # Save configurations before starting the schedule
        self.save_config()

        # Schedule daily check at this time
        schedule.every().day.at(schedule_time.strip()).do(self.master.after, 0, self.check_virustotal)

        # Set schedule_running to True
        self.schedule_running = True

        # Update console window
        self.console_output.write('Schedule started\n')

        # Start schedule thread
        self.schedule_thread = threading.Thread(target=self.run_schedule, daemon=True)
        self.schedule_thread.start()

    def run_schedule(self):
        while self.schedule_running:
            schedule.run_pending()
            time.sleep(1)

        # Update console window
        self.console_output.write('Schedule stopped\n')

    def stop_schedule(self):
        # Enable start button and disable stop button
        self.start_button.configure(state='normal')
        self.stop_button.configure(state='disabled')

        # Clear schedule
        schedule.clear()

        # Set schedule_running to False
        self.schedule_running = False

    def quit(self):
        sys.stdout = sys.__stdout__
        self.master.destroy()

    def show_graph(self):
        self.threat_checker.start()



if __name__ == '__main__':
    # Create root window
    root = tk.Tk()
    root.iconbitmap('ThreatScraper.ico')
    # Create app instance
    app = App(root)

    # Run the application
    root.mainloop()