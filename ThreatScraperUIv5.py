import sys
import threading
import tkinter as tk
from tkinter import messagebox
import requests
import pandas as pd
import datetime
import schedule
import time
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from multiprocessing import Process


class ConsoleOutput:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

    def set_text_widget_state(self, state):
        self.text_widget.configure(state=state)


class VirusTotalChecker:
    def __init__(self, api_key, sha256_hash, filename, console_output):
        self.api_key = api_key
        self.sha256_hash = sha256_hash
        self.url = f'https://www.virustotal.com/vtapi/v2/file/report?apikey={api_key}&resource={sha256_hash}'
        self.filename = filename
        self.console_output = console_output
        self.figure, self.ax = plt.subplots()  # Create figure and axes
        self.line, = self.ax.plot([], [], label='Threat Level')  # Create an empty line

    def check_virustotal(self):
        # Send API request
        response = requests.get(self.url)

        # Check if the request was successful
        if response.status_code == 200:
            # Convert the response to a pandas DataFrame
            df = pd.json_normalize(response.json())

            wb = load_workbook(self.filename)
            ws = wb.active

            # Find the next empty row
            row = ws.max_row + 1

            # Get the existing column headers with whitespaces
            existing_headers = [str(header.value).strip() if isinstance(header.value, str) else str(header.value) for header in ws[1]]

            # Update the column headers without whitespaces
            headers = [header.strip() for header in existing_headers]

            # Convert the DataFrame to a list of lists
            rows = df.values.tolist()

            # Write the rows to the worksheet
            for row_data in rows:
                ws.append(row_data)

            # Update the column headers if they have changed
            if existing_headers != headers:
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)

            # Save the workbook to the Excel file
            wb.save(self.filename)

            # Print success message
            self.console_output.write(f'Report saved to {self.filename}\n')
        else:
            # Print error message
            self.console_output.write(f'Error: {response.status_code} - {response.reason}\n')

        # Update graph
        try:
            df = pd.read_excel(self.filename)

            # Strip the leading and trailing whitespaces from column headers
            df.columns = [str(column).strip() for column in df.columns]

            # Specify the correct column name that contains the numeric values
            target_column = 'positives'  # Replace with the correct column name

            # Ensure that the target column exists and is numeric
            if target_column in df.columns and pd.api.types.is_numeric_dtype(df[target_column]):
                new_data = df[target_column].iloc[-1]
                self.load_data(self.filename)  # Load data from file
                self.update_graph()  # Update the graph
            else:
                self.console_output.write(f'Error: column "{target_column}" is not present or not numeric in {self.filename}\n')

        except Exception as e:
            self.console_output.write(f'Error updating graph: {str(e)}\n')

    def load_data(self, filename):
        try:
            df = pd.read_excel(filename)
            self.ydata = df['positives'].tolist()
            self.xdata = list(range(1, len(self.ydata) + 1))  # Update x-axis values from 1 to the length of ydata
            self.line.set_data([], [])  # Clear previous data
        except Exception as e:
            self.console_output.write(f'Error loading data: {str(e)}')


    def update_graph(self):
        self.line.set_data(self.xdata, self.ydata)  # Update the line data
        self.ax.relim()  # Recalculate limits
        self.ax.autoscale_view(True, True, True)  # Rescale the view
        plt.xlabel('Scan Number')
        plt.ylabel('Threat Level')
        plt.legend()
        plt.draw()
        plt.pause(0.001)

    def start(self):
        self.load_data(self.filename)
        self.update_graph()


class App:
    def __init__(self, master):
        self.master = master
        master.title('ThreatScraper')
        self.threat_checker = None  # Initialize the VirusTotalChecker object as None



        # Create API key entry
        self.api_key_label = tk.Label(master, text='API Key:')
        self.api_key_label.grid(row=0, column=0)
        self.api_key_entry = tk.Entry(master, width=50)
        self.api_key_entry.grid(row=0, column=1)

        # Create SHA-256 hash entry
        self.sha256_hash_label = tk.Label(master, text='SHA-256 Hash:')
        self.sha256_hash_label.grid(row=1, column=0)
        self.sha256_hash_entry = tk.Entry(master, width=50)
        self.sha256_hash_entry.grid(row=1, column=1)

        # Create filename entry
        self.filename_label = tk.Label(master, text='Filename:')
        self.filename_label.grid(row=2, column=0)
        self.filename_entry = tk.Entry(master, width=50)
        self.filename_entry.grid(row=2, column=1)

        # Create time entry
        self.time_label = tk.Label(master, text='Schedule Times (HH:MM) Separated by comma:')
        self.time_label.grid(row=3, column=0)
        self.time_entry = tk.Entry(master, width=50)
        self.time_entry.grid(row=3, column=1)

        # Create check button
        self.check_button = tk.Button(master, text='Check VirusTotal', command=self.check_virustotal)
        self.check_button.grid(row=4, column=1)

        # Create start button
        self.start_button = tk.Button(master, text='Start Schedule', command=self.start_schedule)
        self.start_button.grid(row=5, column=1)

        # Create stop button
        self.stop_button = tk.Button(master, text='Stop Schedule', command=self.stop_schedule, state='disabled')
        self.stop_button.grid(row=6, column=1)

        # Create console window
        self.console_window = tk.Toplevel()
        self.console_window.title('Console Output')
        self.console_text = tk.Text(self.console_window, state='disabled')
        self.console_text.pack(expand=True, fill='both')
        self.console_output = ConsoleOutput(self.console_text)
        self.console_output.set_text_widget_state('normal')
        sys.stdout = self.console_output

        # Initialize schedule variables
        self.schedule_thread = None
        self.schedule_running = False

        # Start the graph
        self.threat_graph = None

    def check_virustotal(self):
        api_key = self.api_key_entry.get()
        sha256_hash = self.sha256_hash_entry.get()
        filename = self.filename_entry.get()
        
        # If VirusTotalChecker object doesn't exist, create it
        if self.threat_checker is None:
            self.threat_checker = VirusTotalChecker(api_key, sha256_hash, filename, self.console_output)
        self.master.after(0, self.threat_checker.check_virustotal)  # Call check_virustotal() on the existing object


    def start_schedule(self):
        # Get schedule times from entry
        schedule_times = self.time_entry.get().split(',')

        # Disable start button and enable stop button
        self.start_button.configure(state='disabled')
        self.stop_button.configure(state='normal')

        # Loop through schedule times and schedule checks
        for schedule_time in schedule_times:
            # Check if time is valid
            try:
                datetime.datetime.strptime(schedule_time.strip(), '%H:%M')
            except ValueError:
                messagebox.showerror('Invalid Time', f'{schedule_time} is not a valid time. Please enter a valid time in HH:MM format')
                return

            # Schedule daily check at this time
            schedule.every().day.at(schedule_time.strip()).do(self.master.after, 0, self.check_virustotal)

        # Set schedule_running to True
        self.schedule_running = True

        # Update console window
        self.console_output.write('Schedule started\n')

        # Start schedule thread
        self.schedule_thread = threading.Thread(target=self.run_schedule, daemon=True)
        self.schedule_thread.start()

    def run_schedule(self):
        while self.schedule_running:
            schedule.run_pending()
            time.sleep(1)

        # Update console window
        self.console_output.write('Schedule stopped\n')

    def stop_schedule(self):
        # Enable start button and disable stop button
        self.start_button.configure(state='normal')
        self.stop_button.configure(state='disabled')

        # Clear schedule
        schedule.clear()

        # Set schedule_running to False
        self.schedule_running = False

    def quit(self):
        sys.stdout = sys.__stdout__
        self.master.destroy()

    def show_graph(self):
        self.threat_checker.start()


if __name__ == '__main__':
    # Create root window
    root = tk.Tk()

    # Create app instance
    app = App(root)

    # Run the application
    root.mainloop()